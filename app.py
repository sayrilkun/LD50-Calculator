import pandas as pd  # pip install pandas openpyxl
import plotly.express as px  # pip install plotly-express
import streamlit as st  # pip install streamlit
from openpyxl import load_workbook
import numpy as np
from sklearn.linear_model import LinearRegression
from PIL import Image

# emojis: https://www.webfx.com/tools/emoji-cheat-sheet/
st.set_page_config(page_title="Median Lethal Dose Calculator", page_icon=":seedling:", layout="wide")
st.title(":seedling: Median Lethal Dose Calculator")
st.markdown("""---""")

# ---- READ EXCEL ----
@st.cache
def get_data(file,sheet_name,first,last):
    wb = load_workbook(filename=file, 
                   read_only=True,
                   data_only=True)
    ws = wb[sheet_name]

    # Read the cell values into a list of lists
    data_rows = []
    for row in ws[first:last]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    # # Transform into dataframe
    # import pandas as pd
    df = pd.DataFrame(data_rows)
    df['Average'] = df.mean(axis=1)
    df['%Germination'] = (df['Average']/25)*100

    return df


def load_data():
    st.dataframe(df)

    average=np.ceil(df['%Germination'].iloc[-1])
    mortality=100-average

    st.text(f"Average%: {average}")
    st.text(f"%Mortality: {mortality}")

    mort_list.append([x[0],mortality])


def normalize(cg_mort, ex_mort):
    normalized_data = ((ex_mort - cg_mort)/(100-cg_mort))*100
    return abs(np.ceil(normalized_data))

def lr():
    regressor = LinearRegression()
    xx= ld_df['Log Dose'].values.reshape(-1, 1)
    yy = ld_df['Probits'].values.reshape(-1, 1)
    regressor.fit(xx, yy)

    pred = regressor.predict(xx)

    #model performance
    from sklearn.metrics import r2_score, mean_squared_error
    mse = mean_squared_error(yy, pred)
    r2 = r2_score(yy, pred)#Best fit lineplt.scatter(x, y)

    # fig_lr= px.line(
    #     x= xx,
    #     y= pred,
    #     markers = True,
    #     color_discrete_sequence=["#000000"] * len(ld_df),
    #     template="plotly_white",
    # )

    # st.plotly_chart(fig_lr, use_container_width=True)
    st.write(f"**R-Squared : {r2}**" )
    st.text(f"Mean Squared Error : {mse}")
    st.text(f"Y-intercept : {regressor.intercept_}")
    st.text(f"Slope : {regressor.intercept_}")
    
    st.markdown("##")
    st.subheader("🚧 Final Decision")
    st.markdown("##")

    image = Image.open('table.png')

    st.image(image, caption='R-Squared Interpretation')
    st.markdown("##")

    if r2 >= 0.70:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have Very strong positive relationship. Thus, the null hypothesis is rejected. ")
    elif r2 >= 0.40 and r2 <= 0.69:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have Strong positive relationship. Thus, the null hypothesis is rejected. ")
    elif r2 >= 0.30 and r2 <= 0.39:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have Moderate positive relationship. Thus, the null hypothesis is rejected. ")
    elif r2 >= 0.20 and r2 <= 0.29:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have Weak positive relationship. Thus, the null hypothesis is rejected. ")
    elif r2 >= 0.01 and r2 <= 0.19:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have No or negligible relationship. Thus, the null hypothesis is rejected. ")
    elif r2 == 0.00:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have No relationship (zero order corellation). Thus, the null hypothesis is rejected. ")
    elif r2 <= -0.01 and r2 >= -0.19:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have No or negligible relationship. Thus, the null hypothesis is accpeted. ")
    elif r2 <= -0.20 and r2 >= -0.29:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have Weak negative relationship. Thus, the null hypothesis is accepted. ")
    elif r2 <= -0.30 and r2 >= -0.39:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have Moderate negative relationship. Thus, the null hypothesis is acppected. ")
    elif r2 <= -0.40 and r2 >= -0.69:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have Strong negative relationship. Thus, the null hypothesis is accpected. ")
    elif r2 <= -0.70:
        st.info(f"Since the R-Squared value is *{round(r2,2)}*, we conclude that the following data have Very strong negative relationship. Thus, the null hypothesis is accepted. ")

data = pd.read_csv('probit.csv')
pTable_df = pd.DataFrame(data)
    
# ---- SIDEBAR ----
side_left, side_right = st.sidebar.columns(2)

pnri = Image.open('pinri.png')
st.sidebar.image(pnri)
# st.sidebar.header("Philippine Nuclear Reasearch Institute")
    # st.write("Philippine Nuclear Reasearch Institute")


mort_list = []
dose = []


uploaded_file = st.sidebar.file_uploader("Choose a file")
# if uploaded_file is not None:
    # df = get_data_from_excel(uploaded_file)
    # df = get_data(uploaded_file)
sheet_name = st.sidebar.text_input('Sheet name')

control = st.sidebar.text_input(
    'Control Group Range', 
    placeholder='Ex: A1,B4')

# try:
if control != "":
    add_name = "Control Group,"+control
    x = add_name.split(",")
    df = get_data(uploaded_file,sheet_name,x[1],x[2])
    st.subheader(f"🧫 {x[0]}")
    load_data()

# except Exception as e:
#     st.sidebar.error("Invalid Input!")

num_exp  = st.sidebar.text_input('Number of Experimental Group')

# try:
if num_exp != "":
    st.markdown("##")
    st.subheader("☢️ Experimental Group")
    for i in range(int(num_exp)):
        i = st.sidebar.text_input(
            f'{i+1}. Dose and Range of Experimental Group', 
            placeholder='Ex: 350,A1,B4')
        if i != "":
            x = i.split(",")
            df = get_data(uploaded_file,sheet_name,x[1],x[2])
            st.subheader(x[0] + " Gy")
            load_data()

# except Exception as e:
#     st.sidebar.error("Invalid Input!")


if st.sidebar.button("Calculate LD50"):
    # st.subheader("Final Output")
    st.markdown("##")
    st.markdown("""---""")

    st.subheader("✔️ Corrected Mortality")
    st.info('If the mortality rate in control is greater than or equal to 10% proceed with correction of mortality rate by using the equation below.')
    st.markdown("##")

    mort_df = pd.DataFrame(mort_list)
    mort_df.columns =['Treatment', 'Mortality']
    cg_mort = mort_df['Mortality'][0]
    if cg_mort >=10:
        mort_df['Corrected % Mortality'] = normalize(cg_mort,mort_df['Mortality'])
        mort_df['Corrected % Mortality'][0] = "N/A"
    else:
        mort_df['Corrected % Mortality'] = "N/A"

    st.dataframe(mort_df)
    mort_df['Corrected % Mortality'][0] = 0

    st.markdown("##")
    st.subheader("🧬 Probit Analysis")

    probit_df = pd.DataFrame(mort_df)
    probit_df.columns = ['Dose (Gy)', 'Mortality', 'Corrected % Mortality']
    
    probit_df["Dose (Gy)"] = pd.to_numeric(probit_df["Dose (Gy)"].loc[1:], downcast="float")
    
    probit_df['Log Dose']= round(np.log10(probit_df['Dose (Gy)'].loc[1:]),2)

# Referring to Probit Table
    if cg_mort >=10:
        b=probit_df['Corrected % Mortality'].values
    else:
        b=probit_df['Mortality'].values

    probit_val = []
    for i in b:
        probit_val.append(pTable_df['Value'][i])

    probit_df['Probits']= probit_val
    
    # probit_df['Log Dose']= probit_df['Log Dose'].apply(lambda x:round(x,2))
    probit_df['Corrected % Mortality'][0] = 'N/A'
    probit_df['Probits'][0] = 'N/A'
    probit_df['Dose (Gy)'][0] = 'Control Group'


    st.dataframe(probit_df)


    st.markdown("##")
    st.subheader("💉 Median Lethal Dose")


    left_column, right_column = st.columns(2)

    p_val = probit_df['Probits'].loc[1:].values
    l_val = probit_df['Log Dose'].loc[1:].values
    data = {'Log Dose': l_val, 'Probits': p_val}
    ld_df = pd.DataFrame(data)

    rows_count = len(df.index)

    # find LD50

    x1 = ld_df['Log Dose'][0]
    x2 = ld_df['Log Dose'].iloc[-1]

    y1 = ld_df['Probits'][0]
    y2 = ld_df['Probits'].iloc[-1]

    m  = (y1-y2)/(x1-x2)
    b  = (x1*y2 - x2*y1)/(x1-x2)

    # x= round((5-b)/m,2)
    x= round((5-b)/m,2)
    xe = np.ceil(10**2.64)

    with left_column:
        st.dataframe(ld_df)
    with right_column:
        st.subheader(F"LD50= {x}")
        st.subheader(F"{xe} Gy in {rows_count} Days")


    # with right_column:
    fig_ld_50 = px.line(
        ld_df,
        x= "Log Dose",
        y= "Probits",
        markers = True,
        color_discrete_sequence=["#000000"] * len(ld_df),
        template="plotly_white",

    )
    fig_ld_50.update_xaxes(showgrid=False, zeroline=False)
    fig_ld_50.update_yaxes(showgrid=False, zeroline=False)
    fig_ld_50.update_layout(plot_bgcolor="#FFFFFF")
    st.plotly_chart(fig_ld_50, use_container_width=True)

        

    st.subheader("📈 Linear Regression Analysis")
    st.markdown("##")
    st.info('''Hypotheses:

        A. Null: 
        The variability in the mortality of seeds is not explained by the radiation dose.
    B. Alternative: 
        The variability in the mortality of seeds is explained by the radiation dose.
    ''')
    st.markdown("##")
    st.info('''Decision Rule:

        If the R-Squared value is positively significantly correlated, reject null hypothesis.
    ''')
    st.markdown("##")
  
    lr()


# ---- HIDE STREAMLIT STYLE ----
hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            footer:after {
                content:'Made by Research Interns from National University Manila BSc Computer Engineering'; 
                visibility: visible;
                display: block;
                position: relative;
                #background-color: red;
                padding: 5px;
                top: 1px;
            }
            # header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)
