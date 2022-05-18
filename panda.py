import pandas as pd  # pip install pandas openpyxl
import plotly.express as px  # pip install plotly-express
import streamlit as st  # pip install streamlit
from openpyxl import load_workbook

wb = load_workbook(filename=r"C:\Users\luna\Documents\Python\rs.xlsx")
ws = wb['Sheet2']
a= 'C7'
b= 'E17'
# Read the cell values into a list of lists
data_rows = []
for row in ws[a:b]:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

# # Transform into dataframe
# import pandas as pd
df = pd.DataFrame(data_rows)
df['mean'] = df.mean(1)
print(df)
